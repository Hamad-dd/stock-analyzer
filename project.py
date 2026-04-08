import os
import tempfile
import numpy as np
import pandas as pd
import yfinance as yf
import streamlit as st
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split

st.set_page_config(page_title="Smart Market Analyzer", page_icon="📈", layout="wide")


# =========================
# HELPERS
# =========================
def is_valid_ticker(ticker: str) -> bool:
    if not ticker:
        return False
    ticker = ticker.strip().upper()
    allowed = set("ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789.-")
    return all(ch in allowed for ch in ticker)


def safe_float(value):
    try:
        if value is None:
            return None
        return float(value)
    except Exception:
        return None


def format_large_number(num):
    if num is None:
        return "N/A"
    try:
        num = float(num)
        if abs(num) >= 1_000_000_000_000:
            return f"{num / 1_000_000_000_000:.2f}T"
        elif abs(num) >= 1_000_000_000:
            return f"{num / 1_000_000_000:.2f}B"
        elif abs(num) >= 1_000_000:
            return f"{num / 1_000_000:.2f}M"
        elif abs(num) >= 1_000:
            return f"{num / 1_000:.2f}K"
        return f"{num:.2f}"
    except Exception:
        return str(num)


def get_asset_ticker(asset_type, stock_symbol="AAPL"):
    if asset_type == "Stock":
        return stock_symbol.strip().upper()
    elif asset_type == "Gold":
        return "GC=F"
    elif asset_type == "Oil":
        return "CL=F"
    return stock_symbol.strip().upper()


def get_asset_display_name(asset_type, basic, fallback_ticker):
    if asset_type == "Stock":
        return basic.get("Company Name", fallback_ticker)
    return basic.get("Asset Name", asset_type)


def get_file_prefix(asset_type, ticker):
    if asset_type == "Gold":
        return "gold"
    elif asset_type == "Oil":
        return "oil"
    return ticker


def recommendation_badge(recommendation):
    rec = recommendation.lower()
    if "buy" in rec:
        color = "#1f9d55"
        bg = "rgba(31,157,85,0.15)"
    elif "hold" in rec:
        color = "#d4a72c"
        bg = "rgba(212,167,44,0.15)"
    else:
        color = "#d64545"
        bg = "rgba(214,69,69,0.15)"

    return f"""
    <div style="
        padding:10px 14px;
        border-radius:14px;
        background:{bg};
        color:{color};
        font-weight:700;
        display:inline-block;
        border:1px solid rgba(255,255,255,0.08);
    ">
        {recommendation}
    </div>
    """


# =========================
# EVALUATION
# =========================
def evaluate_metric(metric, value):
    if value is None or value == "":
        return "N/A"

    metric_lower = metric.lower()

    if any(x in metric_lower for x in ["company name", "asset name", "sector", "industry", "country", "comment", "entry"]):
        return "Info"

    if "trend" in metric_lower:
        return "Good" if "up" in str(value).lower() else "Bad"

    if any(x in metric_lower for x in ["support", "resistance", "stop loss", "take profit", "market cap", "average volume", "current price", "52 week"]):
        return "Info"

    val = safe_float(value)
    if val is None:
        return "Info"

    if "p/e" in metric_lower:
        if val < 0:
            return "Bad"
        elif val <= 15:
            return "Good"
        elif val <= 30:
            return "Mid"
        else:
            return "Bad"

    if "eps" in metric_lower:
        return "Good" if val > 0 else "Bad"

    if "dividend yield" in metric_lower:
        if val > 0.02:
            return "Good"
        elif val > 0.005:
            return "Mid"
        else:
            return "Bad"

    if "revenue growth" in metric_lower:
        if val > 15:
            return "Good"
        elif val > 5:
            return "Mid"
        else:
            return "Bad"

    if "debt/equity" in metric_lower:
        if val < 1:
            return "Good"
        elif val < 2:
            return "Mid"
        else:
            return "Bad"

    if "beta" in metric_lower:
        if val < 1:
            return "Good"
        elif val <= 1.5:
            return "Mid"
        else:
            return "Bad"

    if "rsi" in metric_lower:
        if val < 30:
            return "Good (Oversold)"
        elif 30 <= val <= 60:
            return "Mid"
        elif 60 < val <= 70:
            return "Good"
        else:
            return "Bad (Overbought)"

    if "volatility" in metric_lower:
        if val < 20:
            return "Good"
        elif val < 35:
            return "Mid"
        else:
            return "Bad"

    if "macd" in metric_lower:
        return "Good" if val > 0 else "Bad"

    if "ai confidence" in metric_lower:
        if val >= 75:
            return "Good"
        elif val >= 60:
            return "Mid"
        else:
            return "Bad"

    return "Info"


def explain_metric(metric):
    metric = metric.lower()
    meanings = {
        "p/e": "Price-to-Earnings ratio",
        "eps": "Earnings Per Share",
        "dividend": "Dividend yield return",
        "revenue growth": "Year-over-year revenue growth",
        "debt/equity": "Financial leverage ratio",
        "rsi": "Relative Strength Index",
        "ma50": "50-day moving average",
        "ma200": "200-day moving average",
        "trend": "Overall price direction",
        "support": "Possible buying floor",
        "resistance": "Possible selling ceiling",
        "stop loss": "Level to limit loss",
        "take profit": "Level to lock profit",
        "entry": "Suggested buy range",
        "market cap": "Total company market value",
        "sector": "Business sector",
        "industry": "Industry classification",
        "average volume": "Average daily traded amount",
        "current price": "Latest market price",
        "beta": "Volatility versus market",
        "volatility": "Price fluctuation level",
        "macd": "Momentum indicator",
        "52 week high": "Highest price in last year",
        "52 week low": "Lowest price in last year",
        "net income": "Profit after expenses",
        "long term debt": "Long-term obligations",
        "cash": "Cash on balance sheet",
        "total equity": "Shareholders' equity",
        "risk level": "Overall trade risk",
        "asset name": "Name of the selected market asset",
        "ai prediction": "Machine learning prediction for next price direction",
        "ai confidence": "Confidence score from the machine learning model",
        "ai model": "The machine learning model used for prediction",
    }
    for key, val in meanings.items():
        if key in metric:
            return val
    return "No explanation available"


# =========================
# ANALYSIS
# =========================
def get_fundamentals(ticker, asset_type="Stock"):
    tk_obj = yf.Ticker(ticker)
    info = tk_obj.info

    def safe_get(k):
        return info.get(k, None)

    if asset_type in ["Gold", "Oil"]:
        default_name = "Gold" if asset_type == "Gold" else "Oil"
        basic = {
            "Asset Name": safe_get("shortName") or default_name,
            "Current Price": safe_get("regularMarketPrice") or safe_get("previousClose"),
            "52 Week High": safe_get("fiftyTwoWeekHigh"),
            "52 Week Low": safe_get("fiftyTwoWeekLow"),
            "Average Volume": safe_get("averageVolume"),
        }
        fin = {}
        return basic, fin

    basic = {
        "Company Name": safe_get("shortName") or ticker,
        "Sector": safe_get("sector"),
        "Industry": safe_get("industry"),
        "Country": safe_get("country"),
        "Current Price": safe_get("regularMarketPrice") or safe_get("previousClose"),
        "52 Week High": safe_get("fiftyTwoWeekHigh"),
        "52 Week Low": safe_get("fiftyTwoWeekLow"),
        "Market Cap": safe_get("marketCap"),
        "EPS": safe_get("trailingEps"),
        "P/E Ratio": safe_get("trailingPE"),
        "Dividend Yield": safe_get("dividendYield"),
        "Average Volume": safe_get("averageVolume"),
        "Beta": safe_get("beta"),
    }

    fin = {}

    try:
        bs = tk_obj.balance_sheet
        if not bs.empty:
            debt = bs.loc["Long Term Debt"].iloc[0] if "Long Term Debt" in bs.index else None
            equity = bs.loc["Total Stockholder Equity"].iloc[0] if "Total Stockholder Equity" in bs.index else None
            cash = bs.loc["Cash"].iloc[0] if "Cash" in bs.index else None

            fin["Long Term Debt"] = debt
            fin["Total Equity"] = equity
            fin["Cash"] = cash

            if debt is not None and equity not in [None, 0]:
                fin["Debt/Equity"] = round(debt / equity, 2)
    except Exception:
        pass

    try:
        fs = tk_obj.financials
        if not fs.empty:
            if "Total Revenue" in fs.index:
                revs = fs.loc["Total Revenue"]
                if len(revs) >= 2 and revs.iloc[1] not in [0, None]:
                    growth = (revs.iloc[0] - revs.iloc[1]) / revs.iloc[1]
                    fin["Revenue Growth (YoY) %"] = round(growth * 100, 2)

            if "Net Income" in fs.index:
                ni = fs.loc["Net Income"]
                if len(ni) >= 1:
                    fin["Net Income"] = ni.iloc[0]
    except Exception:
        pass

    return basic, fin


def compute_technical(ticker):
    tk_obj = yf.Ticker(ticker)
    df = tk_obj.history(period="5y")

    if df.empty:
        raise ValueError("No data found for this asset.")

    df["MA50"] = df["Close"].rolling(window=50).mean()
    df["MA200"] = df["Close"].rolling(window=200).mean()

    delta = df["Close"].diff()
    up = delta.clip(lower=0)
    down = -delta.clip(upper=0)
    roll_up = up.rolling(14).mean()
    roll_down = down.rolling(14).mean()
    rs = roll_up / roll_down
    df["RSI"] = 100 - (100 / (1 + rs))

    ema12 = df["Close"].ewm(span=12, adjust=False).mean()
    ema26 = df["Close"].ewm(span=26, adjust=False).mean()
    df["MACD"] = ema12 - ema26
    df["Signal"] = df["MACD"].ewm(span=9, adjust=False).mean()

    df["Daily Return"] = df["Close"].pct_change()
    volatility = df["Daily Return"].std() * np.sqrt(252) * 100

    latest = df.iloc[-1]
    support = df["Low"].tail(30).min()
    resistance = df["High"].tail(30).max()
    trend = "Uptrend" if latest["MA50"] > latest["MA200"] else "Downtrend"

    tech = {
        "Current Price": round(latest["Close"], 2),
        "MA50": round(latest["MA50"], 2) if not pd.isna(latest["MA50"]) else None,
        "MA200": round(latest["MA200"], 2) if not pd.isna(latest["MA200"]) else None,
        "RSI(14)": round(latest["RSI"], 2) if not pd.isna(latest["RSI"]) else None,
        "MACD": round(latest["MACD"], 2) if not pd.isna(latest["MACD"]) else None,
        "Signal Line": round(latest["Signal"], 2) if not pd.isna(latest["Signal"]) else None,
        "Volatility %": round(volatility, 2) if not pd.isna(volatility) else None,
        "Support (30 days)": round(support, 2),
        "Resistance (30 days)": round(resistance, 2),
        "Trend": trend,
    }

    return tech, df


def suggest_trade(tech):
    price = tech["Current Price"]
    support = tech["Support (30 days)"]
    resistance = tech["Resistance (30 days)"]
    rsi = tech.get("RSI(14)")
    trend = tech.get("Trend", "Unknown")

    if price <= support * 1.05 and trend == "Uptrend":
        entry_zone = f"{support:.2f} - {support * 1.05:.2f}"
        suggestion = "Possible buy opportunity near support in an uptrend."
        risk = "Moderate"
    elif price <= support * 1.05:
        entry_zone = f"{support:.2f} - {support * 1.05:.2f}"
        suggestion = "Near support, but trend is weak. Wait for confirmation."
        risk = "High"
    else:
        entry_zone = f"Around {price:.2f}"
        suggestion = "Not near support. Better to wait for a better entry."
        risk = "Moderate"

    stop_loss = support * 0.97
    take_profit = resistance * 0.95

    if rsi is not None and rsi > 70:
        suggestion += " RSI suggests overbought conditions."
        risk = "High"
    elif rsi is not None and rsi < 30:
        suggestion += " RSI suggests oversold conditions."

    return {
        "Suggested Entry Zone": entry_zone,
        "Support Level": round(support, 2),
        "Resistance Level": round(resistance, 2),
        "Stop Loss": round(stop_loss, 2),
        "Take Profit": round(take_profit, 2),
        "Risk Level": risk,
        "Comment": suggestion,
    }


# =========================
# AI PREDICTION
# =========================
def ai_prediction(df):
    data = df.copy()

    required_cols = ["Close", "MA50", "MA200", "RSI", "MACD"]
    for col in required_cols:
        if col not in data.columns:
            raise ValueError(f"Missing required feature: {col}")

    data["Target"] = (data["Close"].shift(-1) > data["Close"]).astype(int)
    data = data.dropna(subset=required_cols + ["Target"])

    if len(data) < 60:
        return "Not enough data", 0.0, "Random Forest"

    X = data[required_cols]
    y = data["Target"]

    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=0.2, random_state=42, shuffle=True
    )

    model = RandomForestClassifier(n_estimators=100, random_state=42)
    model.fit(X_train, y_train)

    latest = X.iloc[[-1]]
    pred = model.predict(latest)[0]
    prob = model.predict_proba(latest)[0]

    direction = "Up 📈" if pred == 1 else "Down 📉"
    confidence = round(float(max(prob)) * 100, 2)

    return direction, confidence, "Random Forest"


def show_rsi_alert(rsi):
    if rsi is None:
        return
    if rsi > 70:
        st.warning("⚠️ RSI indicates overbought conditions.")
    elif rsi < 30:
        st.success("🔥 RSI indicates oversold conditions.")
    else:
        st.info("ℹ RSI is in a normal range.")


# =========================
# SCORING
# =========================
def metric_to_score(rating):
    if rating.startswith("Good"):
        return 2
    if rating == "Mid":
        return 1
    if rating.startswith("Bad"):
        return 0
    return None


def calculate_overall_score(basic, fin, tech):
    all_data = {}
    all_data.update(basic)
    all_data.update(fin)
    all_data.update(tech)

    scored = []
    for k, v in all_data.items():
        rating = evaluate_metric(k, v)
        score = metric_to_score(rating)
        if score is not None:
            scored.append(score)

    if not scored:
        return 0, "Insufficient Data"

    avg = sum(scored) / len(scored)
    if avg >= 1.5:
        rec = "Buy / Positive Outlook"
    elif avg >= 0.9:
        rec = "Hold / Neutral Outlook"
    else:
        rec = "Avoid / Weak Outlook"

    return round(avg, 2), rec


# =========================
# TABLES
# =========================
def make_df_with_rating(data_dict):
    rows = []
    for metric, value in data_dict.items():
        display_value = format_large_number(value) if isinstance(value, (int, float, np.integer, np.floating)) else value
        rows.append({
            "Metric": metric,
            "Value": display_value if value is not None else "N/A",
            "Rating": evaluate_metric(metric, value),
            "Meaning": explain_metric(metric),
        })
    return pd.DataFrame(rows)


# =========================
# CHARTS
# =========================
def create_chart_figure(asset_label, df, trade):
    fig, ax = plt.subplots(figsize=(12, 5.5))
    ax.plot(df.index, df["Close"], label="Close", linewidth=1.8)
    ax.plot(df.index, df["MA50"], label="MA50", linestyle="--")
    ax.plot(df.index, df["MA200"], label="MA200", linestyle="--")
    ax.axhline(trade["Support Level"], linestyle=":", label="Support")
    ax.axhline(trade["Resistance Level"], linestyle=":", label="Resistance")
    ax.axhline(trade["Stop Loss"], linestyle="--", label="Stop Loss")
    ax.axhline(trade["Take Profit"], linestyle="--", label="Take Profit")
    ax.set_title(f"{asset_label} Analysis Chart")
    ax.set_xlabel("Date")
    ax.set_ylabel("Price")
    ax.grid(True, alpha=0.25)
    ax.legend(loc="best")
    fig.tight_layout()
    return fig


def create_comparison_chart(label1, df1, label2, df2):
    fig, ax = plt.subplots(figsize=(12, 5.5))

    base1 = df1["Close"].dropna()
    base2 = df2["Close"].dropna()

    norm1 = (base1 / base1.iloc[0]) * 100
    norm2 = (base2 / base2.iloc[0]) * 100

    ax.plot(norm1.index, norm1.values, label=label1, linewidth=2)
    ax.plot(norm2.index, norm2.values, label=label2, linewidth=2)

    ax.set_title(f"{label1} vs {label2} Comparison")
    ax.set_xlabel("Date")
    ax.set_ylabel("Normalized Performance (Base = 100)")
    ax.grid(True, alpha=0.25)
    ax.legend(loc="best")
    fig.tight_layout()
    return fig


# =========================
# EXCEL
# =========================
def style_worksheet(ws):
    header_fill = PatternFill("solid", fgColor="14213D")
    header_font = Font(color="FFFFFF", bold=True)
    info_fill = PatternFill("solid", fgColor="D9EAF7")
    good_fill = PatternFill("solid", fgColor="C6EFCE")
    mid_fill = PatternFill("solid", fgColor="FFEB9C")
    bad_fill = PatternFill("solid", fgColor="FFC7CE")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            text = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(text))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)

    headers = [c.value for c in ws[1]]
    if "Rating" in headers:
        rating_col = headers.index("Rating") + 1
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=rating_col)
            value = str(cell.value)
            if value.startswith("Good"):
                cell.fill = good_fill
            elif value == "Mid":
                cell.fill = mid_fill
            elif value.startswith("Bad"):
                cell.fill = bad_fill
            else:
                cell.fill = info_fill


def export_to_excel_bytes(file_prefix, summary_name, basic, fin, tech, trade, score, recommendation, fig, ai_info=None):
    df_basic = make_df_with_rating(basic)
    df_fin = make_df_with_rating(fin)
    df_tech = make_df_with_rating(tech)
    df_trade = make_df_with_rating(trade)

    summary_rows = [
        ["Asset", summary_name],
        ["Current Price", basic.get("Current Price")],
        ["Trend", tech.get("Trend")],
        ["RSI(14)", tech.get("RSI(14)")],
        ["Overall Score", score],
        ["Final Recommendation", recommendation],
    ]

    if ai_info:
        summary_rows.extend([
            ["AI Prediction", ai_info.get("AI Prediction")],
            ["AI Confidence", ai_info.get("AI Confidence")],
            ["AI Model", ai_info.get("AI Model")],
        ])

    summary_df = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    tmp_dir = tempfile.mkdtemp()
    xlsx_path = os.path.join(tmp_dir, f"{file_prefix}_analysis.xlsx")
    chart_path = os.path.join(tmp_dir, f"{file_prefix}_chart.png")
    fig.savefig(chart_path, dpi=150, bbox_inches="tight")

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        df_basic.to_excel(writer, sheet_name="Fundamentals", index=False)
        df_fin.to_excel(writer, sheet_name="Financials", index=False)
        df_tech.to_excel(writer, sheet_name="Technical", index=False)
        df_trade.to_excel(writer, sheet_name="Trade Plan", index=False)

    wb = load_workbook(xlsx_path)
    for name in wb.sheetnames:
        style_worksheet(wb[name])

    ws_chart = wb.create_sheet("Chart")
    img = XLImage(chart_path)
    img.width = 900
    img.height = 450
    ws_chart.add_image(img, "A1")
    wb.save(xlsx_path)

    with open(xlsx_path, "rb") as f:
        data = f.read()
    return data


# =========================
# UI
# =========================
st.markdown(
    """
    <style>
    .stApp {background: linear-gradient(180deg, #0b1020 0%, #121a31 100%);}
    .hero {
        padding: 18px 22px;
        border-radius: 20px;
        background: linear-gradient(135deg, rgba(32,46,84,0.95), rgba(17,24,45,0.95));
        border: 1px solid rgba(255,255,255,0.08);
        box-shadow: 0 10px 30px rgba(0,0,0,0.18);
        margin-bottom: 14px;
    }
    .metric-card {
        padding: 16px;
        border-radius: 18px;
        background: rgba(255,255,255,0.04);
        border: 1px solid rgba(255,255,255,0.08);
    }
    .small-label {color:#a8b4d8;font-size:13px;margin-bottom:4px;}
    .big-value {color:white;font-size:22px;font-weight:700;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <div class="hero">
        <h1 style="margin:0;color:white;">📈 Smart Market Analyzer</h1>
        <p style="color:#a8b4d8;font-size:16px;margin-top:6px;">
            Comprehensive analysis for stocks, gold, and oil with comparison mode and AI prediction.
        </p>
    </div>
    """,
    unsafe_allow_html=True,
)

with st.sidebar:
    st.header("Controls")

    mode = st.radio("Mode", ["Single Analysis", "Comparison"])

    st.subheader("Asset 1")
    asset_type_1 = st.selectbox("Asset Type 1", ["Stock", "Gold", "Oil"], key="asset1")

    if asset_type_1 == "Stock":
        stock_symbol_1 = st.text_input("Stock Symbol 1", value="AAPL").strip().upper()
    else:
        stock_symbol_1 = ""

    if mode == "Comparison":
        st.subheader("Asset 2")
        asset_type_2 = st.selectbox("Asset Type 2", ["Stock", "Gold", "Oil"], key="asset2")

        if asset_type_2 == "Stock":
            stock_symbol_2 = st.text_input("Stock Symbol 2", value="MSFT").strip().upper()
        else:
            stock_symbol_2 = ""
    else:
        asset_type_2 = None
        stock_symbol_2 = ""

    analyze = st.button("Analyze", use_container_width=True)
    st.caption("Examples: Stock = AAPL, TSLA | Gold | Oil")

if analyze:
    ticker1 = get_asset_ticker(asset_type_1, stock_symbol_1)

    if asset_type_1 == "Stock" and not is_valid_ticker(ticker1):
        st.error("Please enter a valid first stock ticker.")
    else:
        try:
            basic1, fin1 = get_fundamentals(ticker1, asset_type=asset_type_1)
            tech1, df1 = compute_technical(ticker1)
            trade1 = suggest_trade(tech1)
            score1, recommendation1 = calculate_overall_score(basic1, fin1, tech1)
            asset_name1 = get_asset_display_name(asset_type_1, basic1, ticker1)
            file_prefix1 = get_file_prefix(asset_type_1, ticker1)

            if mode == "Single Analysis":
                fig1 = create_chart_figure(asset_name1, df1, trade1)
                ai_dir, ai_conf, ai_model = ai_prediction(df1)

                c1, c2, c3, c4 = st.columns(4)
                with c1:
                    st.markdown(
                        f'<div class="metric-card"><div class="small-label">📌 Asset</div><div class="big-value">{asset_name1}</div></div>',
                        unsafe_allow_html=True
                    )
                with c2:
                    st.markdown(
                        f'<div class="metric-card"><div class="small-label">💰 Current Price</div><div class="big-value">{format_large_number(basic1.get("Current Price"))}</div></div>',
                        unsafe_allow_html=True
                    )
                with c3:
                    st.markdown(
                        f'<div class="metric-card"><div class="small-label">📈 Trend</div><div class="big-value">{tech1.get("Trend", "N/A")}</div></div>',
                        unsafe_allow_html=True
                    )
                with c4:
                    st.markdown(
                        f'<div class="metric-card"><div class="small-label">🤖 AI Prediction</div><div class="big-value">{ai_dir}</div></div>',
                        unsafe_allow_html=True
                    )

                st.markdown("### Recommendation")
                st.markdown(recommendation_badge(recommendation1), unsafe_allow_html=True)

                st.markdown("### Overall Score")
                st.progress(min(score1 / 2, 1.0))
                st.write(f"Score: **{score1} / 2.0**")

                show_rsi_alert(tech1.get("RSI(14)"))

                left, right = st.columns([1.5, 1])

                with left:
                    st.subheader("Price Chart")
                    st.pyplot(fig1, use_container_width=True)

                with right:
                    st.subheader("Decision Summary")
                    st.write(f"**Overall Score:** {score1}")
                    st.write(f"**RSI(14):** {tech1.get('RSI(14)')}")
                    st.write(f"**MACD:** {tech1.get('MACD')}")
                    st.write(f"**Risk Level:** {trade1.get('Risk Level')}")
                    st.write(f"**Comment:** {trade1.get('Comment')}")

                    st.subheader("AI Module")
                    st.write(f"**Prediction:** {ai_dir}")
                    st.write(f"**Confidence:** {ai_conf}%")
                    st.write(f"**Model:** {ai_model}")

                tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
                    ["Summary", "Fundamentals", "Financials", "Technical", "Trade Plan", "AI Insights"]
                )

                summary_df = make_df_with_rating({
                    "Asset": asset_name1,
                    "Current Price": basic1.get("Current Price"),
                    "Trend": tech1.get("Trend"),
                    "RSI(14)": tech1.get("RSI(14)"),
                    "Overall Score": score1,
                    "Final Recommendation": recommendation1,
                })

                ai_df = make_df_with_rating({
                    "AI Prediction": ai_dir,
                    "AI Confidence": ai_conf,
                    "AI Model": ai_model,
                })

                with tab1:
                    st.dataframe(summary_df, use_container_width=True, hide_index=True)
                with tab2:
                    st.dataframe(make_df_with_rating(basic1), use_container_width=True, hide_index=True)
                with tab3:
                    st.dataframe(make_df_with_rating(fin1), use_container_width=True, hide_index=True)
                with tab4:
                    st.dataframe(make_df_with_rating(tech1), use_container_width=True, hide_index=True)
                with tab5:
                    st.dataframe(make_df_with_rating(trade1), use_container_width=True, hide_index=True)
                with tab6:
                    st.dataframe(ai_df, use_container_width=True, hide_index=True)

                excel_bytes = export_to_excel_bytes(
                    file_prefix=file_prefix1,
                    summary_name=asset_name1,
                    basic=basic1,
                    fin=fin1,
                    tech=tech1,
                    trade=trade1,
                    score=score1,
                    recommendation=recommendation1,
                    fig=fig1,
                    ai_info={
                        "AI Prediction": ai_dir,
                        "AI Confidence": ai_conf,
                        "AI Model": ai_model,
                    }
                )

                st.download_button(
                    "Download Excel Report",
                    data=excel_bytes,
                    file_name=f"{file_prefix1}_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            else:
                ticker2 = get_asset_ticker(asset_type_2, stock_symbol_2)

                if asset_type_2 == "Stock" and not is_valid_ticker(ticker2):
                    st.error("Please enter a valid second stock ticker.")
                else:
                    basic2, fin2 = get_fundamentals(ticker2, asset_type=asset_type_2)
                    tech2, df2 = compute_technical(ticker2)
                    trade2 = suggest_trade(tech2)
                    score2, recommendation2 = calculate_overall_score(basic2, fin2, tech2)
                    asset_name2 = get_asset_display_name(asset_type_2, basic2, ticker2)

                    st.subheader("Comparison Overview")

                    a, b = st.columns(2)
                    with a:
                        st.markdown(
                            f'<div class="metric-card"><div class="small-label">Asset 1</div><div class="big-value">{asset_name1}</div></div>',
                            unsafe_allow_html=True
                        )
                        st.write(f"**Price:** {format_large_number(basic1.get('Current Price'))}")
                        st.write(f"**Trend:** {tech1.get('Trend')}")
                        st.write(f"**RSI:** {tech1.get('RSI(14)')}")
                        st.markdown(recommendation_badge(recommendation1), unsafe_allow_html=True)
                        st.progress(min(score1 / 2, 1.0))
                        st.write(f"Score: **{score1} / 2.0**")

                    with b:
                        st.markdown(
                            f'<div class="metric-card"><div class="small-label">Asset 2</div><div class="big-value">{asset_name2}</div></div>',
                            unsafe_allow_html=True
                        )
                        st.write(f"**Price:** {format_large_number(basic2.get('Current Price'))}")
                        st.write(f"**Trend:** {tech2.get('Trend')}")
                        st.write(f"**RSI:** {tech2.get('RSI(14)')}")
                        st.markdown(recommendation_badge(recommendation2), unsafe_allow_html=True)
                        st.progress(min(score2 / 2, 1.0))
                        st.write(f"Score: **{score2} / 2.0**")

                    show_rsi_alert(tech1.get("RSI(14)"))
                    show_rsi_alert(tech2.get("RSI(14)"))

                    fig_compare = create_comparison_chart(asset_name1, df1, asset_name2, df2)
                    st.subheader("Comparison Chart")
                    st.pyplot(fig_compare, use_container_width=True)

                    comp_df = pd.DataFrame([
                        {
                            "Asset": asset_name1,
                            "Current Price": format_large_number(basic1.get("Current Price")),
                            "Trend": tech1.get("Trend"),
                            "RSI(14)": tech1.get("RSI(14)"),
                            "MACD": tech1.get("MACD"),
                            "Score": score1,
                            "Recommendation": recommendation1,
                        },
                        {
                            "Asset": asset_name2,
                            "Current Price": format_large_number(basic2.get("Current Price")),
                            "Trend": tech2.get("Trend"),
                            "RSI(14)": tech2.get("RSI(14)"),
                            "MACD": tech2.get("MACD"),
                            "Score": score2,
                            "Recommendation": recommendation2,
                        },
                    ])

                    st.subheader("Comparison Table")
                    st.dataframe(comp_df, use_container_width=True, hide_index=True)

        except Exception as e:
            st.error(f"Analysis failed: {e}")
else:
    st.info("Choose mode and assets from the sidebar, then press Analyze.")
